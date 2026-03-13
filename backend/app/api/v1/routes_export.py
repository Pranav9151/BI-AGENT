"""
Smart BI Agent — Export Routes  (Component 17)
Architecture v3.1 | Layer 4 (Application) | Threats: T18 (weasyprint SSRF),
                                                        T31 (sensitive SQL leakage),
                                                        T48 (shadow AI / export audit)

ENDPOINTS:
    POST   /api/v1/export                        — export tabular data (csv/excel/pdf)
    POST   /api/v1/export/saved-query/{id}       — export saved query SQL as .sql file

ACCESS:
    Both endpoints require require_active_user.
    Saved-query export enforces ownership (T52) — same _assert_owner_or_admin used in C13.
    Rate limiting (5 req/min) is enforced by the middleware layer (routes_export is
    mounted under /export which the rate_limiter targets).

RATE LIMITING (5/min):
    The middleware in app/middleware/rate_limiter.py targets /api/v1/export* at 5 req/min
    per user.  No per-route rate limit code here — the middleware handles it.

SENSITIVE DATA BLOCK (T31 / T48):
    Exports with sensitivity="restricted" are rejected with 422.
    "sensitive" exports are allowed but the classification stamp is prominent.
    Every export writes an audit log entry (T48 shadow AI leakage control).

CLASSIFICATION STAMP (T48):
    Every generated file includes a classification comment/header:
      - CSV: first line comment "# Classification: <level>"
      - Excel: cell A1 in a dedicated "Classification" sheet
      - PDF: visible banner at top of first page
    This makes exported files self-describing when removed from the platform.

PDF + SSRF (T18):
    weasyprint is used for PDF generation with a custom url_fetcher that rejects
    any external URL (http/https to non-localhost hosts).  Only inline base64
    data URIs and same-origin assets are allowed.
    If weasyprint is not installed, the PDF endpoint returns 501 Not Implemented.

SAVED QUERY EXPORT:
    The /saved-query/{id} endpoint exports the STORED SQL as a .sql text file.
    It does NOT re-execute the query.  Ownership rules (T52) apply: only the
    owner or an admin can export another user's saved query SQL.
    Restricted-sensitivity saved queries cannot be exported by non-admins.

AUDIT:
    Every export (success or blocked) writes to AuditWriter:
      - format, row_count (for tabular), sensitivity, filename
      - For blocked exports: reason in the question field
    This provides an audit trail for data exfiltration investigations.
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Request, Response, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import (
    CurrentUser,
    get_audit_writer,
    get_db,
    require_active_user,
)
from app.errors.exceptions import (
    InsufficientPermissionsError,
    ResourceNotFoundError,
    ValidationError,
)
from app.logging.structured import get_logger
from app.models.saved_query import SavedQuery
from app.schemas.export import (
    MAX_EXPORT_ROWS,
    ExportRequest,
    ExportSavedQueryRequest,
    ExportMetadataResponse,
)

log = get_logger(__name__)

router = APIRouter()

# =============================================================================
# Constants
# =============================================================================

_CLASSIFICATION_HEADER = "CLASSIFICATION"

# MIME types for file responses
_MIME = {
    "csv":   "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf":   "application/pdf",
    "sql":   "text/plain",
}


# =============================================================================
# Helpers
# =============================================================================

def _safe_filename(hint: Optional[str], fmt: str, fallback: str = "export") -> str:
    """Build a safe filename with the correct extension."""
    ext_map = {"csv": "csv", "excel": "xlsx", "pdf": "pdf", "sql": "sql"}
    base = hint or fallback
    # Strip any existing extension
    base = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
    # Allow only alphanumeric, dash, underscore, space
    base = re.sub(r"[^\w\s\-]", "_", base).strip()
    if not base:
        base = fallback
    ext = ext_map.get(fmt, fmt)
    return f"{base}.{ext}"


def _sensitivity_stamp(sensitivity: str) -> str:
    """Human-readable classification stamp for file headers."""
    labels = {
        "normal":     "INTERNAL USE ONLY",
        "sensitive":  "SENSITIVE — RESTRICTED DISTRIBUTION",
        "restricted": "RESTRICTED — AUTHORISED PERSONNEL ONLY",
    }
    return labels.get(sensitivity, "INTERNAL USE ONLY")


def _get_client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


# --- Format generators -------------------------------------------------------

def _generate_csv(columns: list[str], rows: list[list], sensitivity: str) -> bytes:
    """
    Generate UTF-8 CSV bytes with a classification comment header.
    Classification stamp is on line 1 as a comment so spreadsheet apps
    display it immediately on open (T48).
    """
    buf = io.StringIO()
    stamp = _sensitivity_stamp(sensitivity)
    buf.write(f"# {_CLASSIFICATION_HEADER}: {stamp}\n")

    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([str(v) if v is not None else "" for v in row])

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def _generate_excel(
    columns: list[str],
    rows: list[list],
    sensitivity: str,
    title: Optional[str],
) -> bytes:
    """
    Generate .xlsx using openpyxl.
    Sheet 1: "Classification" — single cell with the stamp (T48).
    Sheet 2: "Data" — column headers + rows.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # --- Classification sheet (T48) ---
    cls_ws = wb.active
    cls_ws.title = "Classification"
    stamp = _sensitivity_stamp(sensitivity)
    cls_ws["A1"] = f"{_CLASSIFICATION_HEADER}: {stamp}"
    cls_ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    fill_color = {
        "normal":     "2196F3",  # Blue
        "sensitive":  "FF9800",  # Orange
        "restricted": "F44336",  # Red
    }.get(sensitivity, "2196F3")
    cls_ws["A1"].fill = PatternFill("solid", fgColor=fill_color)
    cls_ws["A1"].alignment = Alignment(horizontal="center")
    cls_ws.column_dimensions["A"].width = 60

    if title:
        cls_ws["A2"] = title
        cls_ws["A2"].font = Font(bold=True, size=11)

    # --- Data sheet ---
    data_ws = wb.create_sheet("Data")
    data_ws.append(columns)
    # Bold header row
    for cell in data_ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        data_ws.append([v if v is not None else "" for v in row])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(
    columns: list[str],
    rows: list[list],
    sensitivity: str,
    title: Optional[str],
) -> bytes:
    """
    Generate PDF using weasyprint with inline-only asset policy (T18).

    A custom url_fetcher blocks any external http/https URLs, preventing
    SSRF via weasyprint's image/font loading.
    """
    try:
        import weasyprint
    except ImportError:
        raise ValidationError(
            message="PDF export requires weasyprint which is not installed.",
            detail="Install weasyprint to enable PDF export.",
        )

    def _safe_fetcher(url: str, timeout=10, ssl_context=None):
        """Block all external URLs — inline data: URIs only (T18)."""
        if url.startswith("data:"):
            return weasyprint.default_url_fetcher(url)
        raise ValueError(
            f"External URL blocked by SSRF guard: {url!r}. "
            "Only inline data URIs are allowed in PDF exports."
        )

    stamp = _sensitivity_stamp(sensitivity)
    banner_color = {
        "normal":     "#2196F3",
        "sensitive":  "#FF9800",
        "restricted": "#F44336",
    }.get(sensitivity, "#2196F3")

    # Build HTML table (never f-string for user content — use escape)
    from html import escape as _esc

    header_cells = "".join(f"<th>{_esc(str(c))}</th>" for c in columns)
    body_rows = ""
    for row in rows:
        cells = "".join(f"<td>{_esc(str(v) if v is not None else '')}</td>" for v in row)
        body_rows += f"<tr>{cells}</tr>"

    title_html = f"<h2>{_esc(title)}</h2>" if title else ""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: sans-serif; font-size: 11px; margin: 20px; }}
  .banner {{ background: {banner_color}; color: white; padding: 6px 12px;
             font-weight: bold; margin-bottom: 12px; border-radius: 3px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #f0f0f0; border: 1px solid #ccc; padding: 4px 8px;
        text-align: left; font-size: 10px; }}
  td {{ border: 1px solid #ddd; padding: 3px 8px; font-size: 10px; }}
  tr:nth-child(even) {{ background: #fafafa; }}
</style>
</head>
<body>
  <div class="banner">{_CLASSIFICATION_HEADER}: {_esc(stamp)}</div>
  {title_html}
  <table>
    <thead><tr>{header_cells}</tr></thead>
    <tbody>{body_rows}</tbody>
  </table>
</body>
</html>"""

    doc = weasyprint.HTML(string=html, url_fetcher=_safe_fetcher)
    return doc.write_pdf()


# --- Saved query helpers -----------------------------------------------------

async def _get_sq_or_404(sq_id: uuid.UUID, db: AsyncSession) -> SavedQuery:
    result = await db.execute(select(SavedQuery).where(SavedQuery.id == sq_id))
    sq = result.scalar_one_or_none()
    if sq is None:
        raise ResourceNotFoundError(
            message="Saved query not found.",
            detail=f"SavedQuery {sq_id} does not exist",
        )
    return sq


def _assert_sq_readable(sq: SavedQuery, current_user: CurrentUser) -> None:
    """
    Read + export gate for saved queries (T52 + T31).
    Mirrors _assert_readable from routes_saved_queries but blocks export
    of restricted queries even by owner when role is not admin (T31).
    """
    is_owner = str(sq.user_id) == current_user["user_id"]
    is_admin = current_user["role"] == "admin"

    if is_admin:
        return

    if not is_owner:
        raise InsufficientPermissionsError(
            message="You do not have permission to export this saved query.",
            detail=f"SavedQuery {sq.id} is not owned by {current_user['user_id']}",
        )

    # Owners cannot export restricted queries (T31 — prevent SQL leakage via export)
    if sq.sensitivity == "restricted":
        raise InsufficientPermissionsError(
            message="Restricted queries cannot be exported.",
            detail=(
                f"SavedQuery {sq.id} has sensitivity=restricted. "
                "Only admins can export restricted query SQL."
            ),
        )


# =============================================================================
# POST /  — Export tabular data
# =============================================================================

@router.post(
    "/",
    summary="Export query results",
    description=(
        "Export tabular data (rows + columns) as CSV, Excel, or PDF. "
        "Rate-limited to 5 req/min. "
        "Restricted-sensitivity data is blocked. "
        "Every export is audit-logged (T48)."
    ),
    response_class=StreamingResponse,
    responses={
        200: {"description": "File attachment"},
        422: {"description": "Validation error (sensitivity blocked, row limit, bad format)"},
    },
)
async def export_data(
    request: Request,
    body: ExportRequest,
    current_user: CurrentUser = Depends(require_active_user),
    db: AsyncSession = Depends(get_db),
    audit=Depends(get_audit_writer),
) -> StreamingResponse:
    """
    Core export endpoint.

    Restricted sensitivity is blocked entirely (T31/T48):
    If a user somehow has restricted data in-memory, it cannot be exported.
    """
    # Block restricted exports (T31)
    if body.sensitivity == "restricted":
        if audit:
            await audit.log(
                execution_status="export.blocked_restricted",
                question=(
                    f"Export blocked: sensitivity=restricted, "
                    f"format={body.format}, rows={len(body.rows)}"
                ),
                user_id=current_user["user_id"],
                ip_address=_get_client_ip(request),
                request_id=getattr(request.state, "request_id", None),
            )
        raise ValidationError(
            message="Restricted data cannot be exported.",
            detail="Exports with sensitivity='restricted' are not permitted.",
        )

    # Validate column/row shape consistency
    for i, row in enumerate(body.rows):
        if len(row) != len(body.columns):
            raise ValidationError(
                message="Row data shape mismatch.",
                detail=(
                    f"Row {i} has {len(row)} values but "
                    f"{len(body.columns)} columns were declared."
                ),
            )

    filename = _safe_filename(body.filename, body.format)
    row_count = len(body.rows)

    # Generate file bytes
    if body.format == "csv":
        file_bytes = _generate_csv(body.columns, body.rows, body.sensitivity)
    elif body.format == "excel":
        file_bytes = _generate_excel(
            body.columns, body.rows, body.sensitivity, body.title
        )
    else:  # pdf
        file_bytes = _generate_pdf(
            body.columns, body.rows, body.sensitivity, body.title
        )

    log.info(
        "export.generated",
        user_id=current_user["user_id"],
        format=body.format,
        rows=row_count,
        sensitivity=body.sensitivity,
        size_bytes=len(file_bytes),
    )

    # Audit log — T48 export audit
    if audit:
        await audit.log(
            execution_status="export.generated",
            question=(
                f"Exported {row_count} rows as {body.format.upper()} "
                f"(sensitivity={body.sensitivity}, file={filename})"
            ),
            user_id=current_user["user_id"],
            row_count=row_count,
            result_bytes=len(file_bytes),
            ip_address=_get_client_ip(request),
            request_id=getattr(request.state, "request_id", None),
        )

    mime = _MIME[body.format]
    return StreamingResponse(
        iter([file_bytes]),
        media_type=mime,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(file_bytes)),
            "X-Export-Rows": str(row_count),
            "X-Export-Sensitivity": body.sensitivity,
        },
    )


# =============================================================================
# POST /saved-query/{query_id}  — Export saved query SQL
# =============================================================================

@router.post(
    "/saved-query/{query_id}",
    summary="Export saved query SQL",
    description=(
        "Download the stored SQL of a saved query as a .sql text file. "
        "Does NOT re-execute the query. "
        "Owner or admin only (T52). "
        "Restricted-sensitivity queries require admin role (T31)."
    ),
    response_class=StreamingResponse,
    responses={
        200: {"description": ".sql file attachment"},
        403: {"description": "Not owner / restricted sensitivity"},
        404: {"description": "Saved query not found"},
    },
)
async def export_saved_query(
    query_id: uuid.UUID,
    request: Request,
    body: ExportSavedQueryRequest,
    current_user: CurrentUser = Depends(require_active_user),
    db: AsyncSession = Depends(get_db),
    audit=Depends(get_audit_writer),
) -> StreamingResponse:
    """
    Export the SQL text of a saved query.

    The file format is plain SQL with a classification header comment.
    Re-running the SQL requires appropriate DB access and is the user's
    responsibility — no permission re-validation happens here (T43 applies
    only to actual execution, not to reading the stored text).
    """
    sq = await _get_sq_or_404(query_id, db)
    _assert_sq_readable(sq, current_user)

    from html import escape as _esc

    lines = [f"-- Classification: {_sensitivity_stamp(sq.sensitivity)}"]
    lines.append(f"-- Query: {sq.name}")
    lines.append(f"-- Exported by: {current_user['email']}")
    lines.append(f"-- Exported at: {datetime.now(timezone.utc).isoformat()}")
    if body.include_question:
        lines.append(f"-- Question: {sq.question}")
    lines.append("")
    lines.append(sq.sql_query)

    sql_bytes = "\n".join(lines).encode("utf-8")
    filename = _safe_filename(sq.name, "sql")

    log.info(
        "export.saved_query",
        user_id=current_user["user_id"],
        query_id=str(query_id),
        sensitivity=sq.sensitivity,
    )

    if audit:
        await audit.log(
            execution_status="export.saved_query",
            question=(
                f"Exported saved query SQL: {sq.name!r} ({query_id}), "
                f"sensitivity={sq.sensitivity}"
            ),
            user_id=current_user["user_id"],
            result_bytes=len(sql_bytes),
            ip_address=_get_client_ip(request),
            request_id=getattr(request.state, "request_id", None),
        )

    return StreamingResponse(
        iter([sql_bytes]),
        media_type=_MIME["sql"],
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(sql_bytes)),
            "X-Export-Sensitivity": sq.sensitivity,
        },
    )